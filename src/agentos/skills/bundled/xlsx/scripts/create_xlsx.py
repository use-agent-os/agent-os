"""Create a `.xlsx` workbook from a JSON spec.

Spec:
    {
      "sheets": [
        {
          "name": "Sales",
          "rows": [["A", "B"], [1, "=B1*2"]],
          "merged": [{"range": "A1:B1"}],
          "freeze": "A2"
        }
      ]
    }
"""

from __future__ import annotations

import argparse
import json
import sys
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook


def _coerce(value: Any) -> Any:
    if isinstance(value, str) and len(value) >= 19 and value[10] == "T":
        try:
            return datetime.fromisoformat(value)
        except ValueError:
            return value
    return value


def build(spec: dict[str, Any]) -> Workbook:
    wb = Workbook()
    default_sheet = wb.active
    sheets = spec.get("sheets") or []
    if not sheets:
        return wb

    for idx, sheet_spec in enumerate(sheets):
        if not isinstance(sheet_spec, dict):
            continue
        if idx == 0:
            ws = default_sheet
            ws.title = str(sheet_spec.get("name") or "Sheet1")
        else:
            ws = wb.create_sheet(title=str(sheet_spec.get("name") or f"Sheet{idx + 1}"))

        for row in sheet_spec.get("rows", []):
            ws.append([_coerce(v) for v in row])

        for merged in sheet_spec.get("merged") or []:
            if isinstance(merged, dict) and "range" in merged:
                ws.merge_cells(str(merged["range"]))

        freeze = sheet_spec.get("freeze")
        if isinstance(freeze, str) and freeze:
            ws.freeze_panes = freeze

    return wb


def _parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Create a .xlsx from a JSON spec.")
    parser.add_argument("spec", type=Path)
    parser.add_argument("--out", type=Path, required=True)
    return parser.parse_args()


def main() -> int:
    args = _parse_args()
    if not args.spec.is_file():
        print(f"error: spec {args.spec} not found", file=sys.stderr)
        return 2
    spec = json.loads(args.spec.read_text(encoding="utf-8"))
    wb = build(spec)
    args.out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(args.out))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
