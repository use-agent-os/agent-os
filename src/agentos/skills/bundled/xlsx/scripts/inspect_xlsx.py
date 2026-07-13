"""Dump `.xlsx` workbook structure as JSON for LLM consumption."""

from __future__ import annotations

import argparse
import json
import sys
from datetime import date, datetime, time
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import TYPE_BOOL, TYPE_FORMULA, TYPE_NUMERIC, TYPE_STRING


def _serialize(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time):
        return value.isoformat()
    return value


def inspect(path: Path, data_only: bool) -> dict[str, Any]:
    wb = load_workbook(filename=str(path), data_only=data_only)
    sheets: list[dict[str, Any]] = []
    for ws in wb.worksheets:
        rows: list[list[dict[str, Any]]] = []
        for row in ws.iter_rows():
            rows.append(
                [
                    {
                        "value": _serialize(cell.value),
                        "type": cell.data_type,
                    }
                    for cell in row
                ]
            )
        merged = sorted(str(r) for r in ws.merged_cells.ranges)
        sheets.append(
            {
                "name": ws.title,
                "max_row": ws.max_row,
                "max_col": ws.max_column,
                "rows": rows,
                "merged": merged,
                "freeze": ws.freeze_panes or "",
            }
        )
    return {
        "sheets": sheets,
        "data_types": {
            "n": "numeric",
            "s": TYPE_STRING,
            "d": "datetime",
            "f": TYPE_FORMULA,
            "b": TYPE_BOOL,
            "n_alias": TYPE_NUMERIC,
        },
    }


def _parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Dump .xlsx structure as JSON.")
    parser.add_argument("path", type=Path, help="Path to a .xlsx workbook")
    parser.add_argument(
        "--data-only",
        action="store_true",
        help="Return cached computed values instead of formula expressions",
    )
    parser.add_argument("--out", type=Path, default=None)
    return parser.parse_args()


def main() -> int:
    args = _parse_args()
    if not args.path.is_file():
        print(f"error: {args.path} not found", file=sys.stderr)
        return 2
    payload = inspect(args.path, args.data_only)
    text = json.dumps(payload, ensure_ascii=False, indent=2, default=str)
    if args.out is not None:
        args.out.write_text(text, encoding="utf-8")
    else:
        print(text)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
