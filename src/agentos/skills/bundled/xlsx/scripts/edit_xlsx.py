"""Apply cell edits to an existing `.xlsx`.

Operations:
    {"op": "set_cell", "sheet": "Q3", "row": 1, "col": 1, "value": "..."}
    {"op": "set_cell", "sheet": "Q3", "row": 2, "col": 2, "value": "=SUM(B3:B10)"}
    {"op": "set_cell", "sheet": "Q3", "row": 3, "col": 3, "value": "=hello", "as_text": true}
    {"op": "rename_sheet", "old": "Sheet1", "new": "Summary"}
    {"op": "merge_cells", "sheet": "Q3", "range": "A1:C1"}
"""

from __future__ import annotations

import argparse
import json
import sys
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def _coerce(value: Any, as_text: bool) -> Any:
    if as_text and isinstance(value, str) and value.startswith("="):
        return "'" + value
    if isinstance(value, str) and len(value) >= 19 and value[10] == "T":
        try:
            return datetime.fromisoformat(value)
        except ValueError:
            return value
    return value


def apply_ops(wb: Any, ops: list[dict[str, Any]]) -> int:
    applied = 0
    for op in ops:
        if not isinstance(op, dict):
            continue
        kind = op.get("op")
        if kind == "set_cell":
            sheet_name = op.get("sheet")
            row = op.get("row")
            col = op.get("col")
            value = op.get("value")
            if sheet_name not in wb.sheetnames or row is None or col is None:
                continue
            ws = wb[sheet_name]
            ws.cell(row=int(row), column=int(col), value=_coerce(value, bool(op.get("as_text"))))
            applied += 1
        elif kind == "rename_sheet":
            old = op.get("old")
            new = op.get("new")
            if old in wb.sheetnames and isinstance(new, str):
                wb[old].title = new
                applied += 1
        elif kind == "merge_cells":
            sheet_name = op.get("sheet")
            rng = op.get("range")
            if sheet_name in wb.sheetnames and isinstance(rng, str):
                wb[sheet_name].merge_cells(rng)
                applied += 1
    return applied


def _parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Edit an .xlsx via JSON op list.")
    parser.add_argument("input", type=Path)
    parser.add_argument("ops", type=Path)
    parser.add_argument("--out", type=Path, required=True)
    return parser.parse_args()


def main() -> int:
    args = _parse_args()
    if not args.input.is_file():
        print(f"error: input {args.input} not found", file=sys.stderr)
        return 2
    if not args.ops.is_file():
        print(f"error: ops {args.ops} not found", file=sys.stderr)
        return 2
    raw = json.loads(args.ops.read_text(encoding="utf-8"))
    ops = raw if isinstance(raw, list) else []
    wb = load_workbook(filename=str(args.input))
    applied = apply_ops(wb, ops)
    args.out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(args.out))
    print(json.dumps({"applied": applied}, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
