"""xlsx skill — load, eligibility, and create→inspect→edit round-trip."""

from __future__ import annotations

import sys
from pathlib import Path

import pytest

from agentos.skills.eligibility import EligibilityContext, check_eligibility
from agentos.skills.loader import SkillLoader

ROOT = Path(__file__).resolve().parents[1]
BUNDLED = ROOT / "src" / "agentos" / "skills" / "bundled"
SCRIPTS = BUNDLED / "xlsx" / "scripts"


def _spec() -> object:
    return SkillLoader(bundled_dir=BUNDLED).get_by_name("xlsx")


def test_skill_loads() -> None:
    spec = _spec()
    assert spec is not None
    assert spec.name == "xlsx"
    assert spec.metadata is not None


def test_eligibility_with_python(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setattr(
        "agentos.skills.eligibility.shutil.which",
        lambda name: "/usr/bin/python3" if name in {"python", "python3"} else None,
    )
    spec = _spec()
    assert spec is not None
    assert check_eligibility(spec, EligibilityContext.auto())


def test_round_trip_with_formula_and_merge(tmp_path: Path) -> None:
    sys.path.insert(0, str(SCRIPTS))
    try:
        import create_xlsx  # type: ignore[import-not-found]
        import edit_xlsx  # type: ignore[import-not-found]
        import inspect_xlsx  # type: ignore[import-not-found]
    finally:
        sys.path.pop(0)

    spec = {
        "sheets": [
            {
                "name": "Sales",
                "rows": [
                    ["Region", "Revenue"],
                    ["NA", 1_200_000],
                    ["EU", 850_000],
                    ["Total", "=SUM(B2:B3)"],
                ],
                "merged": [{"range": "A1:B1"}],
                "freeze": "A2",
            }
        ]
    }
    src = tmp_path / "book.xlsx"
    create_xlsx.build(spec).save(str(src))
    assert src.exists()

    inspected = inspect_xlsx.inspect(src, data_only=False)
    sheet = next(s for s in inspected["sheets"] if s["name"] == "Sales")
    assert sheet["max_row"] == 4
    assert sheet["max_col"] == 2
    assert any("A1:B1" in r for r in sheet["merged"])
    assert sheet["freeze"] == "A2"

    last_row = sheet["rows"][3]
    assert last_row[1]["type"] == "f"
    assert last_row[1]["value"] == "=SUM(B2:B3)"

    from openpyxl import load_workbook

    wb = load_workbook(str(src))
    edit_xlsx.apply_ops(
        wb,
        [
            {"op": "set_cell", "sheet": "Sales", "row": 2, "col": 1, "value": "Americas"},
            {"op": "rename_sheet", "old": "Sales", "new": "Q3"},
        ],
    )
    out = tmp_path / "out.xlsx"
    wb.save(str(out))

    re_inspected = inspect_xlsx.inspect(out, data_only=False)
    sheet_q3 = next(s for s in re_inspected["sheets"] if s["name"] == "Q3")
    assert sheet_q3["rows"][1][0]["value"] == "Americas"


def test_text_escapes_formula(tmp_path: Path) -> None:
    sys.path.insert(0, str(SCRIPTS))
    try:
        import create_xlsx  # type: ignore[import-not-found]
        import edit_xlsx  # type: ignore[import-not-found]
        import inspect_xlsx  # type: ignore[import-not-found]
    finally:
        sys.path.pop(0)

    src = tmp_path / "book.xlsx"
    create_xlsx.build({"sheets": [{"name": "S", "rows": [["a"]]}]}).save(str(src))

    from openpyxl import load_workbook

    wb = load_workbook(str(src))
    edit_xlsx.apply_ops(
        wb,
        [
            {
                "op": "set_cell",
                "sheet": "S",
                "row": 2,
                "col": 1,
                "value": "=hello",
                "as_text": True,
            },
        ],
    )
    out = tmp_path / "out.xlsx"
    wb.save(str(out))

    inspected = inspect_xlsx.inspect(out, data_only=False)
    sheet = inspected["sheets"][0]
    cell_value = sheet["rows"][1][0]["value"]
    assert isinstance(cell_value, str)
    assert cell_value.lstrip("'") == "=hello"
